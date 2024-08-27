import csv
from hmac import new
import json
import re
import openpyxl as pl
# from database import sql
import os
from tkinter import *
#execl表格输出
def table_create():
    excel_olt = pl.load_workbook('xx双上联改造资源申请表.xlsx',data_only=True)
    sheet = excel_olt.active

    table = []
    max_row = sheet.max_row

    for i in range(max_row):
        table.append([])

    for i,row in enumerate(sheet):
        for cell in row:
            table[i].append(cell.value)

    return table
#对文档内容处理，分成不同表格
def columnFromLightLine(waypass:list):
    #👎分成四列数据：字典暂时使用下，未来优化
    pattern = {'电路编码':'21Z.*$','电路路由':'.*/.*/*$|与','光路编码':'21G.*$','光路路由':'---.*$'}
    plist = [i for i in pattern.values()]
    print(plist)
    results = []

    def patternToStr(waypass,pattern):
        lines = []
        for line in waypass:
            matchojb = re.search(pattern,line)
            # print(matchojb)
            if matchojb != None:
                lines.append(line)
        return lines
    #👎输出结果应使用字典，现在为string
    for patten in plist:
        result = patternToStr(waypass,patten)
        results.append(result)

    return results

#获取一二级分光器信息及对应关系
def getODNupANDdown(jtext,filename):#filename包含路径
    jdata:json = json.loads(jtext)
    qinq = pl.Workbook()
    sheet = qinq.active
    oneeqp = []                     #一级分光器列表
    oneeqp_order = []               #一级分光器对应行数,例[['a',1],['b',2]]
    flag_2EQP = 1                   #判断是否有第二台分光器,1 有,0 无
    try:
        result = jdata[0]['TWOEQP']
    except KeyError:
        flag_2EQP = 0#确认有无二级分光器
        print('no sercond onu exit!')
    
    # 显示局站名称和站点编码
    # 
    # 暂未实现

    for i,data in enumerate(jdata):

        firstpoint = data['ONEEQP']

        if flag_2EQP == 1:
            sheet.cell(i+2,3).value = data['TWOEQP']
            print("{:20}".format(" "),"分光器2:",data['TWOEQP'])
        
        if firstpoint not in oneeqp:
            sheet.cell(i+2,2).value = firstpoint
            oneeqp.append(firstpoint)
            oneeqp_order.append([firstpoint,i+2])

    qinq.save(filename+'(ONU)'+'.xlsx')
    # return sheet

#获取onu上联
def getOltUpofOnu(jtext,filename):

    jdata:json = json.loads(jtext)[0]
    qinq = pl.Workbook()
    sheet = qinq.active
    print(jdata)
    oltName = jdata['OLTNO']
    oltPort = jdata['OLTPORTNO']

    sheet.cell(1,4).value = oltName
    sheet.cell(1,5).value = oltPort

    qinq.save(filename=filename+'(OLT)'+'.xlsx')
    return sheet



#合并ONU sheets包含路径+文件名（未解决olt对应分光器问题）
def mergeONU(sheets:list,sdir:str):
    olt = []
    for i in sheets:
        print(i)
        sheet = pl.load_workbook(i)
        sh = sheet.active
        n = 2
        while(1):   
            olt.append([sh.cell(n,2).value,sh.cell(n,3).value])
            n = n+1
            if sh.cell(n,2).value == None and sh.cell(n,3).value == None:
                # sheet.close()
                break
    
    p = pl.Workbook()
    s = p.active
    for num,i in enumerate(olt):
        s.cell(num+2,1).value = i[0]
        s.cell(num+2,2).value = i[1]
    p.save(sdir)
    p.close()

#输出全部全程路由
def totalRoute(edata:list,sdir:str):
    p = pl.Workbook()
    s = p.active
    num = 1
    for i in edata:
        s.cell(num,1).value = i
        num = num+1
    p.save(sdir+"totalRoute.xlsx")
    p.close()


#re正则匹配
def searchString(patternL,line):

    matchojb = re.search(patternL,line)
    
    return matchojb

class dia:
    eidCode = []
    eidStr  = ""

    def __init__(self) -> None:
        self.f = open(".\database\dia.csv","r")
        self.xl = csv.reader(self.f,delimiter=",")
        pass

    def fNameOfEid(self,fname):
        '''名称找电路'''
        xls = pl.Workbook()
        newSH = xls.active
        i  = 1
        j  = 0
        for row in self.xl:
                sval = row[4]
                if sval == fname:
                    j = j+1
                    newSH.cell(j,1).value = row[4]
                    newSH.cell(j,2).value = row[1]
                    newSH.cell(j,3).value = row[10]
                    newSH.cell(j,4).value = row[11]
                    self.eidCode.append(row[1])
        print(self.eidCode)
        xls.save(fname+"—dia.xlsx")
        xls.close()
        self.f.close()

    def feidofName(self,ls:list):
        xls = pl.Workbook()
        newSH = xls.active
        for row in newSH:
            break
        pass



