from database import sql
import openpyxl as pl


#根据olt上联端口查询电路编码
def findNumberByRID(RID,filename):
    eNid = sql.lite.selectOltByRID(RID)
    olt = pl.load_workbook(filename=filename)
    sheet = olt.active
    sheet.cell(1,6).value = eNid

    olt.save(filename=filename)

#根据olt上联端口+上联名称查询电路编码
def findNumberByOLTPort(oname,oport,filename):
    try:
        eNid = sql.lite.selectOltByPort(oname,oport)
    except IndexError:
        return 0
    if eNid == []:
        return 0
    print(filename)
    olt = pl.load_workbook(filename=filename)
    sheet = olt.active
    sheet.cell(1,6).value = eNid[0]

    olt.save(filename)
